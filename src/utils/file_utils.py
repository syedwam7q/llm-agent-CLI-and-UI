"""
File utility functions for the LLM Agent.
Created by: Syed Wamiq
"""
import json
import csv
from pathlib import Path
from typing import Optional, List, Dict, Any
import sys

# Ensure proper path setup
current_dir = Path(__file__).parent
src_dir = current_dir.parent
sys.path.insert(0, str(src_dir))

from config import settings

# Optional imports for file processing
try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


def find_file(filename: str) -> Optional[Path]:
    """
    Find a file in multiple possible locations.
    
    Args:
        filename: The filename to search for
        
    Returns:
        Path to the file if found, None otherwise
    """
    path = Path(filename)
    
    # If it's already an absolute path and exists, return it
    if path.is_absolute() and path.exists():
        return path
    
    # Search locations in order of preference
    search_locations = [
        settings.data_dir / "uploads",  # Uploaded files
        settings.data_dir,              # Data directory
        settings.project_root,          # Project root
        Path.cwd(),                     # Current working directory
    ]
    
    for location in search_locations:
        candidate = location / path
        if candidate.exists():
            return candidate
    
    return None


def get_file_search_paths() -> List[Path]:
    """
    Get all possible file search paths.
    
    Returns:
        List of paths where files are searched
    """
    return [
        settings.data_dir / "uploads",
        settings.data_dir,
        settings.project_root,
        Path.cwd(),
    ]


def ensure_upload_directory() -> Path:
    """
    Ensure the upload directory exists and return its path.
    
    Returns:
        Path to the upload directory
    """
    upload_dir = settings.data_dir / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    return upload_dir


def get_file_info(file_path: Path) -> dict:
    """
    Get information about a file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Dictionary with file information
    """
    if not file_path.exists():
        return {"exists": False}
    
    stat = file_path.stat()
    return {
        "exists": True,
        "size": stat.st_size,
        "size_mb": round(stat.st_size / (1024 * 1024), 2),
        "extension": file_path.suffix.lower(),
        "name": file_path.name,
        "absolute_path": str(file_path.absolute()),
        "relative_to_project": str(file_path.relative_to(settings.project_root)) if file_path.is_relative_to(settings.project_root) else None,
    }


def is_supported_file_type(filename: str) -> bool:
    """
    Check if a file type is supported.
    
    Args:
        filename: The filename to check
        
    Returns:
        True if the file type is supported
    """
    supported_extensions = {'.txt', '.json', '.csv', '.pdf', '.docx', '.xlsx'}
    return Path(filename).suffix.lower() in supported_extensions


def format_file_size(size_bytes: int) -> str:
    """
    Format file size in human-readable format.
    
    Args:
        size_bytes: Size in bytes
        
    Returns:
        Formatted size string
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"


def process_uploaded_file(file_path: str) -> Optional[str]:
    """
    Process an uploaded file and extract its content.
    
    Args:
        file_path: Path to the uploaded file
        
    Returns:
        Extracted content as string, or None if processing failed
    """
    path = Path(file_path)
    
    if not path.exists():
        return None
    
    extension = path.suffix.lower()
    
    try:
        if extension == '.txt':
            return _process_text_file(path)
        elif extension == '.json':
            return _process_json_file(path)
        elif extension == '.csv':
            return _process_csv_file(path)
        elif extension == '.pdf' and HAS_PDF:
            return _process_pdf_file(path)
        elif extension == '.docx' and HAS_DOCX:
            return _process_docx_file(path)
        elif extension == '.xlsx' and HAS_XLSX:
            return _process_xlsx_file(path)
        else:
            return f"File type {extension} is not supported or required libraries are not installed."
    
    except Exception as e:
        return f"Error processing file: {str(e)}"


def _process_text_file(path: Path) -> str:
    """Process a text file."""
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()
    return f"Text file content:\n\n{content}"


def _process_json_file(path: Path) -> str:
    """Process a JSON file."""
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return f"JSON file content:\n\n{json.dumps(data, indent=2)}"


def _process_csv_file(path: Path) -> str:
    """Process a CSV file."""
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    if not rows:
        return "CSV file is empty."
    
    # Show first few rows
    preview_rows = rows[:10]  # Show first 10 rows
    content = "CSV file content (first 10 rows):\n\n"
    
    for i, row in enumerate(preview_rows):
        content += f"Row {i+1}: {', '.join(row)}\n"
    
    if len(rows) > 10:
        content += f"\n... and {len(rows) - 10} more rows"
    
    return content


def _process_pdf_file(path: Path) -> str:
    """Process a PDF file."""
    if not HAS_PDF:
        return "PDF processing requires PyPDF2 library."
    
    with open(path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        text = ""
        
        for page_num, page in enumerate(reader.pages):
            text += f"\n--- Page {page_num + 1} ---\n"
            text += page.extract_text()
    
    return f"PDF file content:\n{text}"


def _process_docx_file(path: Path) -> str:
    """Process a DOCX file."""
    if not HAS_DOCX:
        return "DOCX processing requires python-docx library."
    
    doc = Document(path)
    text = ""
    
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    
    return f"DOCX file content:\n\n{text}"


def _process_xlsx_file(path: Path) -> str:
    """Process an XLSX file."""
    if not HAS_XLSX:
        return "XLSX processing requires openpyxl library."
    
    workbook = openpyxl.load_workbook(path)
    content = "XLSX file content:\n\n"
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        content += f"Sheet: {sheet_name}\n"
        
        # Get first 10 rows
        for row_num, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
            content += f"Row {row_num}: {', '.join(str(cell) if cell is not None else '' for cell in row)}\n"
        
        if sheet.max_row > 10:
            content += f"... and {sheet.max_row - 10} more rows\n"
        
        content += "\n"
    
    return content


class FileProcessor:
    """Async file processor for web UI."""
    
    def __init__(self):
        self.supported_types = {'.txt', '.json', '.csv', '.pdf', '.docx', '.xlsx'}
    
    async def process_file(self, file_path: Path) -> str:
        """
        Process a file asynchronously and extract its content.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Extracted content as string
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extension = file_path.suffix.lower()
        
        if extension not in self.supported_types:
            raise ValueError(f"Unsupported file type: {extension}")
        
        try:
            if extension == '.txt':
                return await self._process_text_file_async(file_path)
            elif extension == '.json':
                return await self._process_json_file_async(file_path)
            elif extension == '.csv':
                return await self._process_csv_file_async(file_path)
            elif extension == '.pdf':
                return self._process_pdf_file_sync(file_path)
            elif extension == '.docx':
                return self._process_docx_file_sync(file_path)
            elif extension == '.xlsx':
                return self._process_xlsx_file_sync(file_path)
        except Exception as e:
            raise RuntimeError(f"Error processing file: {str(e)}")
    
    async def _process_text_file_async(self, path: Path) -> str:
        """Process a text file asynchronously."""
        import aiofiles
        async with aiofiles.open(path, 'r', encoding='utf-8') as f:
            content = await f.read()
        return f"Text file content:\n\n{content}"
    
    async def _process_json_file_async(self, path: Path) -> str:
        """Process a JSON file asynchronously."""
        import aiofiles
        async with aiofiles.open(path, 'r', encoding='utf-8') as f:
            content = await f.read()
            data = json.loads(content)
        return f"JSON file content:\n\n{json.dumps(data, indent=2)}"
    
    async def _process_csv_file_async(self, path: Path) -> str:
        """Process a CSV file asynchronously."""
        import aiofiles
        async with aiofiles.open(path, 'r', encoding='utf-8') as f:
            content = await f.read()
            reader = csv.reader(content.splitlines())
            rows = list(reader)
        
        if not rows:
            return "CSV file is empty."
        
        # Show first few rows
        preview_rows = rows[:10]
        content = "CSV file content (first 10 rows):\n\n"
        
        for i, row in enumerate(preview_rows):
            content += f"Row {i+1}: {', '.join(row)}\n"
        
        if len(rows) > 10:
            content += f"\n... and {len(rows) - 10} more rows"
        
        return content
    
    def _process_pdf_file_sync(self, path: Path) -> str:
        """Process a PDF file synchronously."""
        return _process_pdf_file(path)
    
    def _process_docx_file_sync(self, path: Path) -> str:
        """Process a DOCX file synchronously."""
        return _process_docx_file(path)
    
    def _process_xlsx_file_sync(self, path: Path) -> str:
        """Process an XLSX file synchronously."""
        return _process_xlsx_file(path)